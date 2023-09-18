

import numpy as np
import math as m
import scipy.stats as stats
from openpyxl import Workbook, load_workbook


class StockOption:
    def __init__(self):
        self._DeleteWorkbookSheets_() # To remove output sheets from previous runs
       # self._GetStockOptionData_()
        self.CalcGreeks = 0


        # using openpyxl functionality
        self.wb_models = load_workbook('stockoptiondata.xlsx')
        self.ws_models = self.wb_models['Stock Option']
        self.S0 = self.ws_models['B4'].value
        self.K = self.ws_models['B5'].value
        self.r = self.ws_models['B6'].value
        self.div = self.ws_models['B7'].value
        self.T = self.ws_models['B8'].value
        self.sigma = self.ws_models['B9'].value
        # Ensure n has at least 1 time step
        self.n = max(1,self.ws_models['B10'].value) 
        self.IsCall = self.ws_models['B9'].value
        self.IsEuropean = self.ws_models['B10'].value
             
    def _DeleteWorkbookSheets_(self): 
        # utility function to remove output sheets from previous runs
        wb = load_workbook('stockoptiondata.xlsx')
        # a sheet providing comparisons for different expiries
        
        if 'Comparisons' in wb.sheetnames:
            ws1 = wb['Comparisons']
        else:
            ws1=wb.create_sheet("Comparisons")

        for row in ws1.iter_rows():
            for cell in row:
                cell.value = None  # You can also set it to an empty string by using cell.value = ''

        ws = wb["Stock Option"]
        ws.freeze_panes = None
        ws1.freeze_panes = None
        ws_active = wb['Stock Option']
        # delete other sheets but Stock Option data sheet
        for sheet in wb.worksheets:
            if sheet.title != 'Stock Option' and sheet.title!='Comparisons':
                wb.remove(sheet) 
        
        for j in range(5,17):
            for i in range(5,9):
                ws.cell(i,j,0.0)
        for j in range(5,17):
            for i in range(12,15):
                ws.cell(i,j,0.0)
            
        wb.save("stockoptiondata.xlsx")
        
class BinomialOption(StockOption):
    def _InitParams_(self):
        self.dt = 0.0 # delta of time step, in years
        self.u = 0.0 # Up move increment
        self.d = 0.0 # Down move increment
        self.pu = 0.0 # Probability of Up move
        self.pd = 0.0 # Probability of Down move
    
    def _CRROptionPrice_(self):
        # Cox, Ross, Rubinstein (1979)
        self._InitParams_()
        self.dt = float(self.T)/float(self.n) # delta of time step, in years
        self.u = np.exp(self.sigma*(np.sqrt(self.dt)))
        self.d = 1 / self.u
        self.pu = (np.exp((self.r- self.div)*self.dt) - self.d) / (self.u - self.d)
        self.pd = 1 - self.pu
        self.model = 'CRR'


        if self.IsCall and self.IsEuropean:
            self.CRREuroCalloptionprice = round(self._CalcOptionPrice_(),4)
        elif not self.IsCall and self.IsEuropean:
            self.CRREuroPutoptionprice = round(self._CalcOptionPrice_(),4)
        elif self.IsCall and not self.IsEuropean:
            self.CRRAmerCalloptionprice = round(self._CalcOptionPrice_(),4)
        else:
            self.CRRAmerPutoptionprice = round(self._CalcOptionPrice_(),4)
     #   if self.CalcGreeks == 0:
    #        self._OutputModelTrees_()
        
    def _LROptionPrice_(self):
        # Leisen, Reimer (1996)
        self._InitParams_()
        temp_n = self.n
        if self.n % 2 == 0: self.n = self.n + 1
        self.dt = float(self.T)/float(self.n) # delta of time step, in years
    
        #Calculate LR up probability and down probability
        d1 = (np.log(self.S0 / self.K) + (((self.r - self.div) + 0.5 * self.sigma**2) * self.T)) \
                    / (self.sigma * np.sqrt(self.T)) # d1
        d2 = d1 - self.sigma * np.sqrt(self.T)
        
        signd1 = -1 if d1 < 0 else 1
        signd2 = -1 if d2 < 0 else 1
        
        # Peizer-Pratt Method 2 
        hzd1_exp_term = (d1 / (self.n + 1 / 3 + (0.1/(self.n+1))))**2 *\
                  (self.n + 1 / 6)
        hzd1 = 0.5 + signd1 / 2.0 * np.sqrt(1 - np.exp(-hzd1_exp_term)) 
        
        hzd2_exp_term = (d2 / (self.n + 1 / 3 + (0.1/(self.n+1))))**2 * \
                   (self.n + 1 / 6)
        hzd2 = 0.5 + signd2 / 2.0 * np.sqrt(1 - np.exp(-hzd2_exp_term))       
        
        pprime = hzd1
        self.pu = hzd2
        self.pd = 1 - self.pu
        
        self.u = np.exp((self.r-self.div) * self.dt) * pprime / self.pu
        self.d = np.exp((self.r-self.div) * self.dt) * (1 - pprime) / \
               (1 - self.pu)
        self.model = 'LR'
        if self.IsCall and self.IsEuropean:
            self.LREuroCalloptionprice = round(self._CalcOptionPrice_(),4)
        elif not self.IsCall and self.IsEuropean:
            self.LREuroPutoptionprice = round(self._CalcOptionPrice_(),4)
        elif self.IsCall and not self.IsEuropean:
            self.LRAmerCalloptionprice = round(self._CalcOptionPrice_(),4)
        else:
            self.LRAmerPutoptionprice = round(self._CalcOptionPrice_(),4)
            
 #       if self.CalcGreeks == 0:
 #           self._OutputModelTrees_()
        self.n = temp_n
        
    def _JROptionPrice_(self):
        # Jarrow, Rudd 
        self._InitParams_()
        self.dt = float(self.T)/float(self.n) # delta of time step, in years
        self.pu = 0.5
        self.pd = 1 - self.pu 
        self.u = np.exp((self.r - self.div - self.sigma**2 * 0.5) * \
                self.dt + self.sigma * np.sqrt(self.dt))
        self.d = np.exp((self.r - self.div - self.sigma**2 * 0.5) * \
               self.dt - self.sigma * np.sqrt(self.dt))
        self.model = 'JR'
        
        if self.IsCall and self.IsEuropean:
            self.JREuroCalloptionprice = round(self._CalcOptionPrice_(),4)
        elif not self.IsCall and self.IsEuropean:
            self.JREuroPutoptionprice = round(self._CalcOptionPrice_(),4)
        elif self.IsCall and not self.IsEuropean:
            self.JRAmerCalloptionprice = round(self._CalcOptionPrice_(),4)
        else:
            self.JRAmerPutoptionprice = round(self._CalcOptionPrice_(),4)
    #    if self.CalcGreeks == 0:
     #       self._OutputModelTrees_()
    
    def _CalcOptionPrice_(self):
        self._StockPriceTree_()
        self._ProbabilityTree_()
        self._PayoffTree_()
        self._BackwardsDiscount_()
        return self.optionprice
    
    def _StockPriceTree_(self):
        self.StockTree = np.zeros((self.n+1,self.n+1))
        for j in range(self.n+1):
            for i in range(j+1):
                self.StockTree[i][j] = self.S0*(self.u**i)*(self.d**(j-i)) 


    def _ProbabilityTree_(self):
        self.ProbTree = np.zeros((self.n+1,self.n+1))
        for j in range(self.n+1):
            for i in range(j+1):
                  self.ProbTree[i][j] = m.factorial(j) / (m.factorial(i) * \
                  m.factorial(j-i))\
                           * self.pu**i * self.pd**(j-i)
      
    def _PayoffTree_(self):
          if self.IsEuropean:
              self._EuropeanPayoffTree_()
          else:
              self._AmericanPayoffTree_()
          
    def _EuropeanPayoffTree_(self):
          #Initialize all payoff nodes to zeros
          self.PayoffTree = np.zeros((self.n+1,self.n+1))
          iopt = 0
          # Get payoffs at terminal nodes at the option maturity
          if self.IsCall:
              iopt = 1
          else:
              iopt = -1
          # Get payoffs at terminal nodes
          for i in range(self.n+1):
              self.PayoffTree[i][self.n] = np.maximum(0, iopt *\
                                 (self.StockTree[i][self.n]-self.K))
              
    def _AmericanPayoffTree_(self):
          #Initialize all payoff nodes to zeros
          self.PayoffTree = np.zeros((self.n+1,self.n+1))
          iopt = 0
          if self.IsCall:
              iopt = 1
          else:
              iopt = -1
          # Get payoffs at each node
          for j in range(self.n+1):
              for i in range(j+1):
                  self.PayoffTree[i][j] = np.maximum(0, iopt * (self.StockTree[i][j]-self.K))
          
    def _BackwardsDiscount_(self):
          if self.IsEuropean:
              TerminalPayoff = 0.0
              for i in range(self.n+1):
                  TerminalPayoff = TerminalPayoff + self.PayoffTree[i][self.n] *\
                         self.ProbTree[i][self.n]
              self.optionprice = TerminalPayoff * np.exp(-self.r * self.T)
          else:
          # Step backward through tree
              if self.IsCall:
                  iopt = 1
              else:
                  iopt = -1
              for j in range(self.n-1,-1,-1): 
                  for i in range(j+1):
                      self.PayoffTree[i][j] = (self.pu * self.PayoffTree[i+1][j+1]\
                             + self.pd * self.PayoffTree[i][j+1]) *\
                               np.exp(-1 * (self.r - self.div) * self.dt)
                      # Use Early exercise price
                      self.PayoffTree[i][j] = np.maximum(iopt *\
                    (self.StockTree[i][j] - self.K), self.PayoffTree[i][j])  
              self.optionprice = self.PayoffTree[0][0]
              
    def _BlackScholesOptionPrice_(self):    
          # Computed values for Black Scholes model
          d1 = (np.log(self.S0 / self.K) + (((self.r - self.div) + 0.5 * \
                             self.sigma**2) * self.T)) \
                      / (self.sigma * np.sqrt(self.T)) # d1
          d2 = (np.log(self.S0 / self.K) + (((self.r - self.div) - 0.5 * \
                              self.sigma**2) * self.T)) \
                      / (self.sigma * np.sqrt(self.T)) # d2
          Nd1 = stats.norm.cdf(d1,0.0,1.0) # N(d1)
          Nd2 = stats.norm.cdf(d2,0.0,1.0) # N(d1)
          Nminusd1 = stats.norm.cdf(-d1,0.0,1.0) # N(-d1)
          Nminusd2 = stats.norm.cdf(-d2,0.0,1.0) # N(-d2)
          
          if self.IsCall: 
              self.BSEuroCalloptionprice = round(self.S0 * np.exp(-self.div * self.T) * Nd1 - \
                  self.K * np.exp(-self.r * self.T) * Nd2,4)
          else:
              self.BSEuroPutoptionprice = round(self.K * np.exp(-self.r * self.T) * Nminusd2 - \
                  self.S0 * np.exp(-self.div * self.T) * Nminusd1,4)
              
    
          
    def _OutputModelPrices_(self):
          # Output model price to Excel file
          self.ws_models['E5'] = self.BSEuroCalloptionprice
          self.ws_models['E6'] = self.CRREuroCalloptionprice
          self.ws_models['E7'] = self.JREuroCalloptionprice
          self.ws_models['E8'] = self.LREuroCalloptionprice
  
  
          self.ws_models['K5'] = self.BSEuroPutoptionprice
          self.ws_models['K6'] = self.CRREuroPutoptionprice
          self.ws_models['K7'] = self.JREuroPutoptionprice
          self.ws_models['K8'] = self.LREuroPutoptionprice
  
  
          self.ws_models['E12'] = self.CRRAmerCalloptionprice
          self.ws_models['E13'] = self.JRAmerCalloptionprice
          self.ws_models['E14'] = self.LRAmerCalloptionprice
          
          self.ws_models['K12'] = self.CRRAmerPutoptionprice
          self.ws_models['K13'] = self.JRAmerPutoptionprice
          self.ws_models['K14'] = self.LRAmerPutoptionprice
      
  
        
    def PriceAll(self):
        self.IsCall = 1
        self.IsEuropean = 1
        
        while self.IsCall and self.IsEuropean:
            self._BlackScholesOptionPrice_()
            self._CRROptionPrice_()
            self._JROptionPrice_()
            self._LROptionPrice_()
            self.IsCall = 0
            
        while not self.IsCall and self.IsEuropean:
            self._BlackScholesOptionPrice_()
            self._CRROptionPrice_()
            self._JROptionPrice_()
            self._LROptionPrice_()
            self.IsCall = 1
            self.IsEuropean = 0
            
        while self.IsCall and not self.IsEuropean:
            self._CRROptionPrice_()
            self._JROptionPrice_()
            self._LROptionPrice_()
            self.IsCall = 0
        
        while not self.IsCall and not self.IsEuropean:
            self._CRROptionPrice_()
            self._JROptionPrice_()
            self._LROptionPrice_()
            break
        self._OutputModelPrices_()
        self.wb_models.save('stockoptiondata.xlsx')
#        self._TheGreeks_()

    def calculate_and_save_comparisons(self, times):

        comparisons_sheet = self.wb_models['Comparisons']
        row_data=['Expiry', 'Am Tree P', 'BS P', 'Am Tree C', 'BS C',]
        comparisons_sheet.append(row_data)
        print("c/p=", self.IsCall, "\t Eu/Am=", self.IsEuropean )
     
        for time in times:
            # Set the time value in the worksheet
            self.ws_models['B8'].value = time
            self.T = time

            row_data= [time]
            self.IsEuropean = False
            self.IsCall = False
            self._CRROptionPrice_()
            self._BlackScholesOptionPrice_()
            row_data.append(self.CRRAmerPutoptionprice)
            row_data.append(self.BSEuroPutoptionprice)

            self.IsCall = True
            self._CRROptionPrice_()
            self._BlackScholesOptionPrice_()
            row_data.append(self.CRRAmerCalloptionprice)
            row_data.append(self.BSEuroCalloptionprice)

        # Append the results to the 'Comparisons' sheet for the current time step
            comparisons_sheet.append(row_data)

        self.wb_models.save('stockoptiondata.xlsx')
    
if __name__ == "__main__":
    
    binomialoption = BinomialOption()
    binomialoption.PriceAll()
    times = np.arange(0.25, 5.25, 0.25)
    binomialoption.calculate_and_save_comparisons(times)
